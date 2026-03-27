import os
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

# .xls (Excel 97-2003)
try:
    import xlrd
    HAS_XLRD = True
except ImportError:
    HAS_XLRD = False

# .xlsx
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


REQUIRED_COLUMNS = ["Дата приказа", "Сотрудник", "Документ", "Должность", "Примечание"]


def _cell_value(cell: Any) -> str:
    if cell is None:
        return ""
    if hasattr(cell, "value"):
        v = cell.value
    else:
        v = cell
    if v is None:
        return ""
    s = str(v).strip()
    return s


def _parse_date(val: Any) -> Optional[datetime]:
    """Парсит дату из ячейки: число Excel, datetime или строка DD.MM.YYYY."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    if hasattr(val, "value"):
        val = val.value
    if isinstance(val, (int, float)):
        # Excel serial date (days since 1899-12-30)
        try:
            from datetime import timedelta
            base = datetime(1899, 12, 30)
            return base + timedelta(days=int(val))
        except (ValueError, OverflowError):
            return None
    if isinstance(val, str):
        val = val.strip()
        for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
            try:
                return datetime.strptime(val[:10], fmt)
            except ValueError:
                continue
    return None


def _find_header_row_and_cols_xls(sheet: "xlrd.sheet.Sheet") -> Tuple[int, Dict[str, int]]:
    cols: Dict[str, int] = {}
    for row_idx in range(min(10, sheet.nrows)):
        for col_idx in range(sheet.ncols):
            cell = sheet.cell(row_idx, col_idx)
            val = _cell_value(cell)
            for need in REQUIRED_COLUMNS:
                if need not in cols and val and need in val:
                    cols[need] = col_idx
        if len(cols) == len(REQUIRED_COLUMNS):
            return row_idx, cols
        cols.clear()
    return 0, cols


def _find_header_row_and_cols_xlsx(ws: "openpyxl.worksheet.worksheet.Worksheet") -> Tuple[int, Dict[str, int]]:
    cols: Dict[str, int] = {}
    for row_idx, row in enumerate(ws.iter_rows(max_row=10), start=1):
        for col_idx, cell in enumerate(row):
            val = _cell_value(cell)
            for need in REQUIRED_COLUMNS:
                if need not in cols and val and need in val:
                    cols[need] = col_idx + 1  # openpyxl 1-based
        if len(cols) == len(REQUIRED_COLUMNS):
            return row_idx, cols
        cols.clear()
    return 1, cols


def _read_sheet_xls(sheet: "xlrd.sheet.Sheet", sheet_name: str) -> List[Dict[str, Any]]:
    header_row, col_map = _find_header_row_and_cols_xls(sheet)
    if len(col_map) < len(REQUIRED_COLUMNS):
        return []
    rows: List[Dict[str, Any]] = []
    for row_idx in range(header_row + 1, sheet.nrows):
        row_data: Dict[str, Any] = {"sheet_name": sheet_name}
        for name, col_idx in col_map.items():
            cell = sheet.cell(row_idx, col_idx)
            raw = cell.value
            if name == "Дата приказа":
                row_data["date"] = _parse_date(raw)
                row_data["date_raw"] = _cell_value(cell)
            else:
                row_data[name] = _cell_value(cell)
        fio = row_data.get("Сотрудник", "").strip()
        if fio:
            rows.append(row_data)
    return rows


def _read_sheet_xlsx(ws: "openpyxl.worksheet.worksheet.Worksheet", sheet_name: str) -> List[Dict[str, Any]]:
    header_row, col_map = _find_header_row_and_cols_xlsx(ws)
    if len(col_map) < len(REQUIRED_COLUMNS):
        return []
    rows = []
    for row in ws.iter_rows(min_row=header_row + 2):
        row_data = {"sheet_name": sheet_name}
        for name, col_idx in col_map.items():
            cell = row[col_idx - 1] if col_idx <= len(row) else None
            raw = cell.value if cell else None
            if name == "Дата приказа":
                row_data["date"] = _parse_date(raw)
                row_data["date_raw"] = _cell_value(cell) if cell else ""
            else:
                row_data[name] = _cell_value(cell) if cell else ""
        fio = row_data.get("Сотрудник", "").strip()
        if fio:
            rows.append(row_data)
    return rows


def load_registry(file_path: str) -> List[Dict[str, Any]]:
    """
    Загружает все листы из файла реестра.
    Возвращает список записей: каждая — dict с ключами
    sheet_name, date, date_raw, Дата приказа, Сотрудник, Документ, Должность, Примечание.
    """
    path = os.path.abspath(file_path)
    if not os.path.isfile(path):
        return []

    ext = os.path.splitext(path)[1].lower()
    all_rows: List[Dict[str, Any]] = []

    if ext == ".xlsx" and HAS_OPENPYXL:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for name in wb.sheetnames:
            ws = wb[name]
            all_rows.extend(_read_sheet_xlsx(ws, name))
        wb.close()
        return all_rows

    if ext == ".xls" and HAS_XLRD:
        wb = xlrd.open_workbook(path)
        for idx in range(wb.nsheets):
            sheet = wb.sheet_by_index(idx)
            name = sheet.name
            all_rows.extend(_read_sheet_xls(sheet, name))
        return all_rows

    if ext == ".xlsx" and not HAS_OPENPYXL:
        raise ImportError("Для .xlsx установите openpyxl")
    if ext == ".xls" and not HAS_XLRD:
        raise ImportError("Для .xls установите xlrd<2: pip install 'xlrd<2'")
    return []
